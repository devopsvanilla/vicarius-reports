import csv
import requests
import urllib3
import json
import os
import sys
import time
import argparse
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from get_ubuntu_oval_status import check_cve

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

MAX_SIZE = 500
REQUEST_DELAY_SECONDS = float(os.environ.get('VRX_VULN_REQUEST_DELAY_SEC', '0'))
MAX_RETRIES = 5
INITIAL_BACKOFF = 2
DASHBOARD_CHART_WIDTH = 16.83  # em cm (aprox. 480 pt)
DASHBOARD_CHART_HEIGHT = 16.83
SCRIPT_DIR = os.path.dirname(__file__)
REPORTS_DIR = os.path.join(SCRIPT_DIR, 'reports')

def load_endpoint_os_info():
    mapping = {}
    filepath = os.path.join(REPORTS_DIR, 'endpoint_so.jsonl')
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    try:
                        data = json.loads(line)
                        mapping[data.get('endpoint')] = {'os': data.get('os', 'N/A'), 'version': data.get('version', 'N/A')}
                    except json.JSONDecodeError:
                        continue
    else:
        print(f"[AVISO] Arquivo {filepath} não encontrado. SO e Versão ficarão indisponíveis.")
    return mapping

def load_dotenv():
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path, 'r') as f:
            for line in f:
                if '=' in line and not line.strip().startswith('#'):
                    k, v = line.strip().split('=', 1)
                    v = v.strip('"').strip("'")
                    if k not in os.environ:
                        os.environ[k] = v

def get_vicarius_credentials():
    base_url = os.environ.get("VICARIUS_BASE_URL")
    api_key = os.environ.get("VICARIUS_API_KEY")

    if not base_url or not api_key:
        print("[ERRO] Credenciais não encontradas (.env ou vars).", file=sys.stderr)
        sys.exit(1)
        
    return base_url.rstrip("/"), api_key

def fetch_all_with_seek(base_url, headers):
    url = f"{base_url}/vicarius-external-data-api/aggregation/searchGroup"
    
    params = {
        'from': 0,
        'size': MAX_SIZE,
        'objectName': 'OrganizationEndpointVulnerabilities',
        'group': 'vulnerabilityId',
        'includeOriginalDoc': 'true',
        'assetCount': 'true',
        'sumLastSubAggregationBuckets': '1',
        'sort': 'aggregationId'
    }

    last_id = None
    page_num = 1
    total_coletado = 0

    print("[INFO] Iniciando extração listando CVEs com paginação SEEK (quebrando limite 10k)...")

    while True:
        current_params = params.copy()
        if last_id is not None:
            current_params['q'] = f"vulnerabilityId<{last_id}"
            
        retries = 0
        backoff = INITIAL_BACKOFF
        success = False

        while retries <= MAX_RETRIES and not success:
            try:
                response = requests.get(url, params=current_params, headers=headers, verify=False)
                
                if response.status_code == 429:
                    retry_after = response.headers.get("X-Rate-Limit-Retry-After-Seconds") or response.headers.get("Retry-After")
                    wait_time = float(retry_after) if retry_after else backoff
                    print(f"[AVISO] Rate Limit 429. Aguardando {wait_time}s...")
                    time.sleep(wait_time)
                    retries += 1
                    backoff *= 2
                    continue
                elif response.status_code >= 500:
                    time.sleep(backoff)
                    retries += 1
                    backoff *= 2
                    continue
                    
                response.raise_for_status()
                data = response.json()
                success = True
            except requests.exceptions.RequestException as e:
                time.sleep(backoff)
                retries += 1
                backoff *= 2

        if not success:
            print("[ERRO] Ocorreu timeout sucessivo.")
            sys.exit(1)

        result_code = data.get('serverResponseResult', {}).get('serverResponseResultCode')
        if result_code != 'SUCCESS':
            print(f"[ERRO] A API retornou erro: {result_code}")
            sys.exit(1)

        items = data.get('serverResponseObject', [])
        
        valid_items = []
        for item in items:
            vulnerability = item.get('aggregationModelAbs', {}).get('organizationEndpointVulnerabilitiesVulnerability')
            if not vulnerability or not vulnerability.get('vulnerabilityEnabled', False):
                continue
                
            endpoint_obj = item.get('aggregationModelAbs', {}).get('organizationEndpointVulnerabilitiesEndpoint', {})
            item['endpointName'] = endpoint_obj.get('endpointName', 'N/A')
            item['endpointId'] = endpoint_obj.get('endpointId', 'N/A')
            
            product_obj = item.get('aggregationModelAbs', {}).get('organizationEndpointVulnerabilitiesProduct', {})
            item['productName'] = product_obj.get('productName', 'N/A')
            
            version_obj = item.get('aggregationModelAbs', {}).get('organizationEndpointVulnerabilitiesVersion', {})
            item['productVersion'] = version_obj.get('versionName', 'N/A')

            kev_action = (vulnerability or {}).get('vulnerabilityCISARequiredAction', '')
            item['kev'] = 'Sim' if kev_action else 'Não'

            valid_items.append(item)

        if not items:
            break

        print(f"[DEBUG] API retornou {len(items)} itens nesta pagina.")

        total_coletado += len(valid_items)
        
        # Envia apenas o lote atual de volta para o iterador usando yield
        yield valid_items
        
        ids = [int(i.get('aggregationModelAbs', {}).get('vulnerabilityId', 0)) for i in items]
        last_id = min(ids)
        print(f"[DEBUG] IDs na resposta - Min: {min(ids)}, Max: {max(ids)}")
        
        print(f"[PROGRESS] {total_coletado} CVEs coletadas (último patch/vuln_id: {last_id})")

        if len(items) < MAX_SIZE:
             break

        page_num += 1
        time.sleep(REQUEST_DELAY_SECONDS)

    print(f"[INFO] Finalizada coleta via SEEK! Total de CVEs consolidadas: {total_coletado}")

def parse_date(date_val):
    if not date_val or date_val == 'N/A':
        return 'N/A'
    if isinstance(date_val, int) or str(date_val).isdigit():
        return datetime.fromtimestamp(int(date_val) / 1000.0)
    try:
        # Tenta interpretar string ISO
        date_str = date_val.replace('Z', '+00:00')
        return datetime.fromisoformat(date_str)
    except Exception:
        return date_val

OVAL_CACHE_FILE = os.path.join(REPORTS_DIR, 'ubuntu_oval_cache.jsonl')

LEGACY_UBUNTU_STATUS_MAP = {
    'not_in_release': 'Not in release',
    'not_affected': 'Not affected',
    'not_affected_or_not_in_release': 'Not affected',
    'vulnerable': 'Vulnerable',
    'fixed': 'Fixed',
    'unknown': 'Needs evaluation',
    'error': 'Needs evaluation',
}


def normalize_oval_status_label(status):
    if status is None:
        return 'Needs evaluation'

    normalized = str(status).strip()
    if not normalized:
        return 'Needs evaluation'

    return LEGACY_UBUNTU_STATUS_MAP.get(normalized.lower(), normalized)


def migrate_oval_cache_file():
    if not os.path.exists(OVAL_CACHE_FILE):
        return False

    migrated_lines = []
    changed = False

    with open(OVAL_CACHE_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.strip():
                continue

            try:
                data = json.loads(line)
            except Exception:
                migrated_lines.append(line.rstrip('\n'))
                continue

            original_status = data.get('status')
            normalized_status = normalize_oval_status_label(original_status)
            if normalized_status != original_status:
                data['status'] = normalized_status
                changed = True

            migrated_lines.append(json.dumps(data, ensure_ascii=False))

    if changed:
        with open(OVAL_CACHE_FILE, 'w', encoding='utf-8') as f:
            for migrated_line in migrated_lines:
                f.write(migrated_line + '\n')

    return changed

def load_oval_cache():
    cache = {}
    if os.path.exists(OVAL_CACHE_FILE):
        migrate_oval_cache_file()
        with open(OVAL_CACHE_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    try:
                        data = json.loads(line)
                        cache[data['key']] = normalize_oval_status_label(data['status'])
                    except Exception:
                        continue
    return cache

def append_to_excel(ws, item, endpoint_mapping, oval_cache=None):
    if oval_cache is None:
        oval_cache = {}
    
    abs_model = item.get('aggregationModelAbs', {})
    vuln = abs_model.get('organizationEndpointVulnerabilitiesVulnerability', {})
    sens_level = vuln.get('vulnerabilitySensitivityLevel', {}) or {}
    threat_level = sens_level.get('sensitivityLevelThreatLevel', {}) or {}
    ext_ref = vuln.get('vulnerabilityExternalReference', {}) or {}
    pub = abs_model.get('organizationEndpointVulnerabilitiesPublisher', {}) or {}
    product = abs_model.get('organizationEndpointVulnerabilitiesProduct', {}) or {}
    prod_ver = abs_model.get('organizationEndpointVulnerabilitiesVersion', {}) or {}
    patch = abs_model.get('organizationEndpointVulnerabilitiesPatch', {}) or {}
    
    endpoint_name = item.get('endpointName', 'N/A')
    os_info = endpoint_mapping.get(endpoint_name, {'os': 'N/A', 'version': 'N/A'})
    
    os_name = os_info['os']
    os_version = os_info['version']
    cve = ext_ref.get('externalReferenceExternalId', 'N/A')
    pkg_name = product.get('productName', 'N/A')
    pkg_version = prod_ver.get('versionName', 'N/A')
    
    oval_status = 'N/A'
    # Verifica se a coluna OS contém "ubuntu" (case-insensitive) para prosseguir com o enriquecimento
    if isinstance(os_name, str) and 'ubuntu' in os_name.lower() and cve != 'N/A' and pkg_name != 'N/A':
        cache_key = f"{os_version}|{cve}|{pkg_name}|{pkg_version}"
        if cache_key in oval_cache:
            oval_status = normalize_oval_status_label(oval_cache[cache_key])
        else:
            print(f"[OVAL] Consultando status de {cve} no pacote {pkg_name}:{pkg_version} para Ubuntu {os_version}...")
            res = check_cve(os_version, cve, {pkg_name: pkg_version})
            oval_status = normalize_oval_status_label(res.get('status', 'Needs evaluation'))
            # Quando a fonte é a API Ubuntu Security, anexa o sub-status (ex.: "Vulnerable, fix deferred")
            if res.get('source') == 'ubuntu_api':
                api_sub = None
                for d in res.get('details', []) or []:
                    if d.get('package', '').lower() == pkg_name.lower():
                        api_sub = d.get('ubuntu_api_status')
                        break
                if api_sub and api_sub not in ('released', 'released-esm', 'not-affected', 'DNE', 'needs-triage'):
                    oval_status = f"{oval_status}, fix {api_sub}"
            oval_cache[cache_key] = oval_status
            
            os.makedirs(os.path.dirname(OVAL_CACHE_FILE), exist_ok=True)
            with open(OVAL_CACHE_FILE, 'a', encoding='utf-8') as cf:
                json.dump({"key": cache_key, "status": oval_status}, cf)
                cf.write('\n')

    kev_action = vuln.get('vulnerabilityCISARequiredAction', '')
    kev = 'Sim' if kev_action else 'Não'

    row = [
        endpoint_name,
        os_info['os'],
        os_info['version'],
        pub.get('publisherName', 'N/A'),
        product.get('productName', 'N/A'),
        prod_ver.get('versionName', 'N/A'),
        vuln.get('vulnerabilitySummary', 'N/A'),
        ext_ref.get('externalReferenceExternalId', 'N/A'),
        sens_level.get('sensitivityLevelName', 'N/A'),
        parse_date(vuln.get('vulnerabilityCreatedAt')),
        parse_date(vuln.get('vulnerabilityUpdatedAt')),
        oval_status,
        kev
    ]
    ws.append(row)


def _write_summary_table(ws, title, start_row, start_col, summary_data, label_header, table_name):
    title_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_fill = PatternFill(fill_type='solid', fgColor='2F75B5')
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    label_col = start_col
    count_col = start_col + 1

    ws.merge_cells(start_row=start_row, start_column=label_col, end_row=start_row, end_column=count_col)
    title_cell = ws.cell(row=start_row, column=label_col, value=title)
    title_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    header_left = ws.cell(row=start_row + 1, column=label_col, value=label_header)
    header_right = ws.cell(row=start_row + 1, column=count_col, value='Quantidade de CVEs únicas')
    for c in (header_left, header_right):
        c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_thin

    sorted_items = sorted(
        ((k, len(v)) for k, v in summary_data.items()),
        key=lambda x: (-x[1], str(x[0]))
    )

    if not sorted_items:
        c1 = ws.cell(row=start_row + 2, column=label_col, value='Sem dados Ubuntu')
        c2 = ws.cell(row=start_row + 2, column=count_col, value=0)
        c1.border = border_thin
        c2.border = border_thin
        c2.number_format = '#,##0'
        top_data_start_row = start_row + 2
        top_data_end_row = start_row + 2
        top_n = 1
    else:
        for idx, (label, count) in enumerate(sorted_items, start=start_row + 2):
            label_cell = ws.cell(row=idx, column=label_col, value=label)
            count_cell = ws.cell(row=idx, column=count_col, value=count)
            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            count_cell.alignment = Alignment(horizontal='center', vertical='center')
            count_cell.number_format = '#,##0'
            label_cell.border = border_thin
            count_cell.border = border_thin

        top_n = min(3, len(sorted_items))
        top_data_start_row = start_row + 2
        top_data_end_row = top_data_start_row + top_n - 1

    table_start_col = get_column_letter(label_col)
    table_end_col = get_column_letter(count_col)
    data_rows = max(1, len(sorted_items))
    table_ref = f"{table_start_col}{start_row + 1}:{table_end_col}{start_row + 1 + data_rows}"
    tab = Table(displayName=table_name, ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tab)

    return {
        'top_n': top_n,
        'top_data_start_row': top_data_start_row,
        'top_data_end_row': top_data_end_row,
        'label_col': label_col,
        'count_col': count_col,
    }


def _add_top3_pie_chart(ws_dashboard, source_ws, title, source_meta, chart_anchor):
    pie = PieChart()
    pie.title = f"{title} (Top {source_meta['top_n']})"
    data = Reference(
        source_ws,
        min_col=source_meta['count_col'],
        min_row=source_meta['top_data_start_row'],
        max_row=source_meta['top_data_end_row']
    )
    labels = Reference(
        source_ws,
        min_col=source_meta['label_col'],
        min_row=source_meta['top_data_start_row'],
        max_row=source_meta['top_data_end_row']
    )
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.style = 10
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showLeaderLines = True
    pie.dataLabels.showCatName = True  # Mostrar apenas o nome da categoria
    pie.dataLabels.showLegendKey = False  # Não mostrar código da legenda
    pie.dataLabels.showSerName = False  # Não mostrar nome da série
    pie.dataLabels.position = 'outEnd'  # Melhor ajuste visual para pizza
    pie.legend = None  # Remove legenda
    pie.height = DASHBOARD_CHART_HEIGHT
    pie.width = DASHBOARD_CHART_WIDTH
    ws_dashboard.add_chart(pie, chart_anchor)


def build_ubuntu_dashboard(workbook, data_ws):
    dashboard_name = 'Dashboard Ubuntu'
    dashboard_tables_name = 'Dashboard Tabelas'

    if dashboard_name in workbook.sheetnames:
        del workbook[dashboard_name]
    if dashboard_tables_name in workbook.sheetnames:
        del workbook[dashboard_tables_name]

    ws_dashboard = workbook.create_sheet(title=dashboard_name)
    ws_tables = workbook.create_sheet(title=dashboard_tables_name)
    ws_tables.sheet_state = 'hidden'
    ws_dashboard.sheet_view.showGridLines = False

    header_to_col = {
        str(data_ws.cell(row=1, column=col).value).strip(): col
        for col in range(1, data_ws.max_column + 1)
        if data_ws.cell(row=1, column=col).value is not None
    }

    # Adiciona a data de geração ao final da tabela 'Vulnerabilidades Ativas'
    if 'Vulnerabilidades Ativas' in workbook.sheetnames:
        ws_ativas = workbook['Vulnerabilidades Ativas']
        last_row = ws_ativas.max_row + 2  # Pula uma linha
        ws_ativas[f"A{last_row}"] = "Data de Geração"
        ws_ativas[f"B{last_row}"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    required_headers = ['Ativo', 'SO', 'Pacote', 'CVE', 'Nível da Vulnerabilidade', 'Status OVAL Ubuntu']
    missing_headers = [h for h in required_headers if h not in header_to_col]
    if missing_headers:
        ws_dashboard['A1'] = 'Não foi possível gerar o dashboard. Colunas ausentes:'
        ws_dashboard['A2'] = ', '.join(missing_headers)
        ws_tables['A1'] = 'Não foi possível gerar as tabelas. Colunas ausentes:'
        ws_tables['A2'] = ', '.join(missing_headers)
        return

    idx_ativo = header_to_col['Ativo']
    idx_so = header_to_col['SO']
    idx_pacote = header_to_col['Pacote']
    idx_cve = header_to_col['CVE']
    idx_nivel = header_to_col['Nível da Vulnerabilidade']
    idx_oval = header_to_col['Status OVAL Ubuntu']

    cves_por_ativo = defaultdict(set)
    cves_por_pacote = defaultdict(set)
    cves_por_nivel = defaultdict(set)
    cves_ubuntu_total = set()
    linhas_ubuntu = 0

    excluded_oval_statuses = {'fixed', 'n/a', 'not affected'}
    for row_idx in range(2, data_ws.max_row + 1):
        so_value = data_ws.cell(row=row_idx, column=idx_so).value
        so_label = str(so_value).strip().lower() if so_value is not None else ''
        # Considera somente SO Ubuntu (case-insensitive)
        if not (so_label == 'ubuntu' or so_label.startswith('ubuntu ')):
            continue

        oval_value = data_ws.cell(row=row_idx, column=idx_oval).value
        oval_status = str(oval_value).strip() if oval_value is not None else ''
        oval_status_normalized = oval_status.lower()
        # Excluir explicitamente status: Fixed, N/A e Not Affected
        if oval_status_normalized in excluded_oval_statuses:
            continue
        # Mantém apenas status relevantes para o dashboard
        if not (
            oval_status_normalized.startswith('vulnerable')
            or oval_status_normalized.startswith('needs evaluation')
            or oval_status_normalized.startswith('fix ignored')
        ):
            continue

        linhas_ubuntu += 1

        cve_value = data_ws.cell(row=row_idx, column=idx_cve).value
        cve = str(cve_value).strip() if cve_value is not None else ''
        if not cve or cve.upper() == 'N/A':
            continue

        cves_ubuntu_total.add(cve)

        ativo_value = data_ws.cell(row=row_idx, column=idx_ativo).value
        pacote_value = data_ws.cell(row=row_idx, column=idx_pacote).value
        nivel_value = data_ws.cell(row=row_idx, column=idx_nivel).value

        ativo = str(ativo_value).strip() if ativo_value not in (None, '') else 'N/A'
        pacote = str(pacote_value).strip() if pacote_value not in (None, '') else 'N/A'
        nivel = str(nivel_value).strip() if nivel_value not in (None, '') else 'N/A'

        cves_por_ativo[ativo].add(cve)
        cves_por_pacote[pacote].add(cve)
        cves_por_nivel[nivel].add(cve)

    ws_dashboard.merge_cells('A1:U2')
    title_cell = ws_dashboard['A1']
    title_cell.value = 'Dashboard Ubuntu — Vulnerabilidades Ativas'
    title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Filtros utilizados
    filters_text = (
        "Filtros aplicados: SO contém 'Ubuntu' (case-insensitive), "
        "Status OVAL Ubuntu ∈ {Needs evaluation, fix ignored, Vulnerable}"
    )
    ws_dashboard['A3'] = filters_text
    ws_dashboard['A3'].font = Font(name='Calibri', size=10, italic=True, color='595959')
    ws_dashboard['A4'] = f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S') }"
    ws_dashboard['A4'].font = Font(name='Calibri', size=10, italic=True, color='595959')

    ws_dashboard['T4'] = 'CVEs únicas Ubuntu'
    ws_dashboard['U4'] = len(cves_ubuntu_total)
    label = ws_dashboard.cell(row=4, column=20)
    value = ws_dashboard.cell(row=4, column=21)
    label.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
    value.font = Font(name='Calibri', size=11, bold=True, color='1F4E78')
    value.number_format = '#,##0'

    ws_dashboard.freeze_panes = 'A6'

    ws_tables.merge_cells('A1:H2')
    tables_title = ws_tables['A1']
    tables_title.value = 'Dashboard Ubuntu — Tabelas de CVEs únicas'
    tables_title.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    tables_title.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    tables_title.alignment = Alignment(horizontal='left', vertical='center')

    tables_subtitle = ws_tables['A3']
    tables_subtitle.value = f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    tables_subtitle.font = Font(name='Calibri', size=10, italic=True, color='595959')

    meta_ativo = _write_summary_table(
        ws_tables,
        title='CVEs únicas por Ativo (Ubuntu)',
        start_row=6,
        start_col=1,
        summary_data=cves_por_ativo,
        label_header='Ativo',
        table_name='DashboardTblAtivo'
    )

    meta_pacote = _write_summary_table(
        ws_tables,
        title='CVEs únicas por Pacote (Ubuntu)',
        start_row=6,
        start_col=4,
        summary_data=cves_por_pacote,
        label_header='Pacote',
        table_name='DashboardTblPacote'
    )

    meta_nivel = _write_summary_table(
        ws_tables,
        title='CVEs únicas por Nível da Vulnerabilidade (Ubuntu)',
        start_row=6,
        start_col=7,
        summary_data=cves_por_nivel,
        label_header='Nível da Vulnerabilidade',
        table_name='DashboardTblNivel'
    )

    # Gráfico 1: linha 7
    _add_top3_pie_chart(
        ws_dashboard,
        ws_tables,
        title='CVEs únicas por Ativo (Ubuntu)',
        source_meta=meta_ativo,
        chart_anchor='A7'
    )

    # Gráfico 2: imediatamente após o primeiro (linha 7 + altura do gráfico)
    # Altura do gráfico em linhas: 16,83cm / 0,35cm ≈ 48 linhas
    # Então, âncora do segundo gráfico: linha 7 + 48 = linha 55
    _add_top3_pie_chart(
        ws_dashboard,
        ws_tables,
        title='CVEs únicas por Pacote (Ubuntu)',
        source_meta=meta_pacote,
        chart_anchor='A55'
    )

    # Gráfico 3: imediatamente após o segundo (linha 55 + 48 = linha 103)
    _add_top3_pie_chart(
        ws_dashboard,
        ws_tables,
        title='CVEs únicas por Nível da Vulnerabilidade (Ubuntu)',
        source_meta=meta_nivel,
        chart_anchor='A103'
    )

    for col in ('A', 'B', 'C', 'D', 'E', 'F'):
        ws_dashboard.column_dimensions[col].width = 4.2
    ws_dashboard.column_dimensions['G'].width = 2.5
    for col in ('H', 'I', 'J', 'K', 'L', 'M'):
        ws_dashboard.column_dimensions[col].width = 4.2
    ws_dashboard.column_dimensions['N'].width = 2.5
    for col in ('O', 'P', 'Q', 'R', 'S', 'T'):
        ws_dashboard.column_dimensions[col].width = 4.2
    ws_dashboard.column_dimensions['U'].width = 12

    ws_dashboard.row_dimensions[1].height = 24
    ws_dashboard.row_dimensions[2].height = 24
    ws_dashboard.row_dimensions[3].height = 18
    ws_dashboard.row_dimensions[4].height = 18
    ws_dashboard.row_dimensions[6].height = 12

    ws_tables.column_dimensions['A'].width = 36
    ws_tables.column_dimensions['B'].width = 18
    ws_tables.column_dimensions['C'].width = 4
    ws_tables.column_dimensions['D'].width = 36
    ws_tables.column_dimensions['E'].width = 18
    ws_tables.column_dimensions['F'].width = 4
    ws_tables.column_dimensions['G'].width = 42
    ws_tables.column_dimensions['H'].width = 18
    ws_tables.freeze_panes = 'A6'

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--force-update', action='store_true', help='Ignora o arquivo local e forca a chamada na API')
    args = parser.parse_args()

    load_dotenv()
    vicarius_base_url, vicarius_api_key = get_vicarius_credentials()

    headers = {
        "Accept": "application/json",
        "Vicarius-Token": vicarius_api_key
    }
    
    endpoint_mapping = load_endpoint_os_info()
    oval_cache = load_oval_cache()

    os.makedirs(REPORTS_DIR, exist_ok=True)
    
    out_file = os.path.join(REPORTS_DIR, 'active_cve.jsonl')
    excel_file = os.path.join(REPORTS_DIR, 'active_cve.xlsx')

    reuse_excel = False
    if not args.force_update and os.path.exists(excel_file):
        resp = input(f"[?] Arquivo '{excel_file}' ja existe. Deseja regerar? [s/N] (N = utilizar o existente): ").strip().lower()
        if resp != 's' and resp != 'y':
            reuse_excel = True

    if reuse_excel:
        print(f"[INFO] Reutilizando '{excel_file}' existente. Pulando coleta e geracao do XLSX.")
        wb = load_workbook(excel_file)
        ws = wb['Vulnerabilidades Ativas'] if 'Vulnerabilidades Ativas' in wb.sheetnames else wb.active
    else:
        fetch_from_api = True
        if not args.force_update and os.path.exists(out_file):
            resp = input(f"[?] Arquivo '{out_file}' ja existe. Deseja regerar consultando a API via Vicarius? [s/N]: ").strip().lower()
            if resp != 's' and resp != 'y':
                fetch_from_api = False

        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerabilidades Ativas"
        columns = [
            'Ativo',
            'SO',
            'Versão do SO',
            'Fornecedor do pacote',
            'Pacote',
            'Versão do Pacote',
            'Descrição da Vulnerabilidade',
            'CVE',
            'Nível da Vulnerabilidade',
            'Data da Criação',
            'Data da Atualização',
            'Status OVAL Ubuntu',
            'KEV'
        ]
        ws.append(columns)

        if fetch_from_api:
            print("[INFO] Iniciando coleta via API (somente endpoints Ubuntu)...")
            with open(out_file, 'w', encoding='utf-8') as f_json:
                for lote_items in fetch_all_with_seek(vicarius_base_url, headers):
                    for item in lote_items:
                        endpoint_name = item.get('endpointName', 'N/A')
                        os_info = endpoint_mapping.get(endpoint_name, {'os': 'N/A', 'version': 'N/A'})
                        if 'ubuntu' not in os_info.get('os', '').lower():
                            continue
                        json.dump(item, f_json)
                        f_json.write('\n')
                        append_to_excel(ws, item, endpoint_mapping, oval_cache)
        else:
            print(f"[INFO] Lendo dados localmente do arquivo '{out_file}' (somente endpoints Ubuntu)...")
            with open(out_file, 'r', encoding='utf-8') as f_json:
                for line in f_json:
                    if line.strip():
                        item = json.loads(line)
                        endpoint_name = item.get('endpointName', 'N/A')
                        os_info = endpoint_mapping.get(endpoint_name, {'os': 'N/A', 'version': 'N/A'})
                        if 'ubuntu' not in os_info.get('os', '').lower():
                            continue
                        append_to_excel(ws, item, endpoint_mapping, oval_cache)

        # Formatar dados como tabela
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row > 1:
            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
            tab = Table(displayName="VulnerabilitiesTable", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Formatando as colunas de data
            for row_idx in range(2, max_row + 1):
                cell_created = ws.cell(row=row_idx, column=10)
                if isinstance(cell_created.value, datetime):
                    cell_created.number_format = 'DD/MM/YYYY HH:MM:SS'

                cell_updated = ws.cell(row=row_idx, column=11)
                if isinstance(cell_updated.value, datetime):
                    cell_updated.number_format = 'DD/MM/YYYY HH:MM:SS'

    build_ubuntu_dashboard(wb, ws)
    wb.save(excel_file)
    print(f"[SUCESSO] Processo concluido! Excel '{excel_file}' atualizado com aba 'Dashboard Ubuntu'.")

    # Gera CSV com o mesmo nome base do XLSX
    csv_file = os.path.splitext(excel_file)[0] + '.csv'
    print(f"[INFO] Gerando CSV '{csv_file}' a partir do XLSX...")
    with open(csv_file, 'w', encoding='utf-8', newline='') as f_csv:
        writer = csv.writer(f_csv)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([
                v.strftime('%d/%m/%Y %H:%M:%S') if isinstance(v, datetime) else v
                for v in row
            ])
    print(f"[SUCESSO] CSV '{csv_file}' gerado.")

if __name__ == "__main__":
    main()

